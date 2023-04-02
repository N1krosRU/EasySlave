from openpyxl import load_workbook
from yattag import Doc, indent
from art import tprint
from pathlib import Path


def check_file(path_to_file):
    if Path(path_to_file).is_file() and Path(path_to_file).suffix == ".xlsx":
        return True
    else:
        return False


def xlsx_to_xml(path_to_file):
    # Parse XLSX
    try:
        wb = load_workbook(filename=path_to_file, data_only=True)  # Open Exel file
        wb.active = 0  # Select first page
        sheet = wb.active

    except:
        return False

    def create_register_list():
        row_list = []
        for row in sheet.iter_rows(min_row=7, max_row=sheet.max_row, values_only=True):
            if row[0] is not None:
                row_list.append(row)
        return row_list

    def create_group_list():
        row_list = []  # Create empty list
        for row in sheet.iter_rows(min_col=8, max_col=8, min_row=7, max_row=sheet.max_row, values_only=True):
            if row[0] is not None:
                row_list.append(row)  # Add all values to list from colum 8 -> Group
        row_set = set(row_list)  # Create set (only unical values)
        row_list = list(row_set)  # Convert in list
        return row_list

    def get_vType(type_str):
        match type_str:
            case "ShortInt":
                return 2
            case "Integer":
                return 3
            case "Float":
                return 4
            case "Double":
                return 5
            case "DateTime":
                return 7
            case "Bool":
                return 11
            case "Word":
                return 18
            case "DWord":
                return 19
            case "Int64":
                return 20
            case "String":
                return 256
            case _:
                return 18

    def get_RegsCount(regscount_str, type_str):
        match type_str:
            case "ShortInt":
                return 0
            case "Integer":
                return 2
            case "Float":
                return 2
            case "Double":
                return 4
            case "DateTime":
                return 4
            case "Bool":
                return 0
            case "Word":
                return 1
            case "DWord":
                return 2
            case "Int64":
                return 4
            case "String":
                return regscount_str
            case _:
                return 1

    def get_DataFormat(format_str):
        match format_str:
            case "0_1":
                return 1
            case "1_0":
                return 2
            case "0_1_2_3":
                return 3
            case "1_0_3_2":
                return 4
            case "2_3_0_1":
                return 5
            case "3_2_1_0":
                return 6
            case "0_1_2_3_4_5_6_7":
                return 7
            case "1_0_3_2_5_4_7_6":
                return 8
            case "2_3_0_1_6_7_4_5":
                return 9
            case "3_2_1_0_7_6_5_4":
                return 10
            case "4_5_6_7_0_1_2_3":
                return 11
            case "5_4_7_6_1_0_3_2":
                return 12
            case "6_7_4_5_2_3_0_1":
                return 13
            case "7_6_5_4_3_2_1_0":
                return 14
            case _:
                return 2

    def get_BitAddr(bit_str, type_str, space_str):
        if type_str == "Bool" and space_str == ("Holding Registers (0x03)" or "Input Registers (0x04)"):
            if 0 <= int(bit_str) <= 15:
                return bit_str
            else:
                return 0
        else:
            return 0

    def get_Description(description_str):
        if description_str is None:
            return "Нет описания"
        else:
            return description_str

    def get_mSpace(space_str):
        match space_str:
            case "Coil (0x01)":
                return 2
            case "Discrete Input (0x02)":
                return 3
            case "Holding Registers (0x03)":
                return 0
            case "Input Registers (0x04)":
                return 1
            case _:
                return 0

    def get_WorkType(worktype_str):
        match worktype_str:
            case "Инкремент":
                return 0
            case "Случайное число":
                return 1
            case "Ручной ввод":
                return 2
            case "Sin":
                return 3
            case "Пила":
                return 4
            case "Скрипт":
                return 5
            case "OPC":
                return 6
            case _:
                return 2

    device_name = sheet["A3"].value
    device_addr = sheet["B3"].value
    device_desc = sheet["C3"].value
    group_list = create_group_list()
    register_row_list = create_register_list()

    # Generate XML tmpl
    try:
        doc, tag, text = Doc().tagtext()
        with tag("server"):
            with tag("Device"):
                with tag("Name"):
                    text(device_name)
                with tag("Number"):
                    text(device_addr)
                with tag("Description"):
                    text(device_desc)
                with tag("NoReply"):
                    text("false")
                with tag("ReplyTimeout"):
                    text("0")
                with tag("ReplyTimeoutTo"):
                    text("0")
                with tag("BitsMerge"):
                    text("false")
                with tag("RegsMerge"):
                    text("false")
                with tag("ProgID"):
                    text("")
                with tag("EnabelRepeat"):
                    text("false")
                with tag("SendTrash"):
                    text("false")
                with tag("InvalidLength"):
                    text("false")
                with tag("Groups"):
                    with tag("Items"):
                        with tag("Name"):
                            text("MainGroup")
                        for group in group_list:
                            with tag("Items"):
                                with tag("Name"):
                                    text(group[0])
                for row in register_row_list:
                    with tag("Register"):
                        with tag("Name"):  # Имя тега
                            text(row[0])
                        with tag("Addr"):  # Адрес
                            text(row[1])
                        with tag("BitAddr"):  # Битовый Адрес
                            text(get_BitAddr(row[2], row[3], row[5]))
                        with tag("vType"):
                            text(get_vType(row[3]))
                        with tag("Description"):  # Описание
                            text(get_Description(row[4]))
                        with tag("mSpace"):
                            text(get_mSpace(row[5]))
                        with tag("ConstGroupName"):  # Константы
                            text("")
                        with tag("Group"):  # Группа
                            text(row[7])
                        with tag("WorkType"):
                            text(get_WorkType(row[8]))
                        with tag("DataFormat"):  # Формат данных
                            text(get_DataFormat(row[9]))
                        with tag("Percent"):  # Флуктуация
                            text("0")
                        with tag("Period"):  # Период
                            text("100")
                        with tag("RegFullName"):  # Полное имя регистра
                            text(f"Сервер.{device_name}.{row[7]}.{row[0]}")
                        with tag("RegsCount"):  # Кол-во регистров
                            text(get_RegsCount(row[13], row[3]))
                        with tag("Value"):  # Значение
                            text(row[15])
                        with tag("DefValue"):  # Значение по умолчанию
                            text(row[15])
        generated_xml = indent(
            doc.getvalue(),
            indentation='    ',
            indent_text=False
        )
        # print(generated_xml)
    except:
        return False

    # Create XML File
    prefix_file_name = path_to_file.split(".")[0]
    sufix_file_name = path_to_file.split(".")[1]
    with open(f"{prefix_file_name}.tmpl", "w", encoding="utf8") as file_obj:
        file_obj.write(generated_xml)

    return True


def main():
    tprint("Easy Slave")
    print("1) Поместите файл таблицы в одну директорию со скриптом")
    print("2) Укажите путь до файла таблицы в формате 'myfile.xlsx': ")
    path_to_book = input()
    if check_file(path_to_book):
        if xlsx_to_xml(path_to_book):
            print("Файл успешно создан!")
        else:
            print("Ошибка создания шаблона, проверьте файл таблицы!")
    else:
        print("Файл несуществует или имеет неверное расширение!")


if __name__ == "__main__":
    main()
